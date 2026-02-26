from django.http import HttpResponse
from io import BytesIO

def export_to_xlsx(queryset, filename, headers, field_mapping):
    """
    Generates an XLSX file from a queryset.
    
    :param queryset: The Django queryset to export.
    :param filename: The name of the file (e.g., 'report.xlsx').
    :param headers: List of column headers (e.g., ['Date', 'User']).
    :param field_mapping: List of model fields or callables corresponding to headers.
    """
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse("Library 'openpyxl' not installed.", status=500)
    except Exception as e:
        return HttpResponse(f"Error loading openpyxl: {str(e)}", status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"
    
    # Write Header
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = openpyxl.styles.Font(bold=True)
        
    # Write Data
    for row_num, obj in enumerate(queryset, 2):
        for col_num, field in enumerate(field_mapping, 1):
            if callable(field):
                val = field(obj)
            else:
                # Handle nested fields (e.g., 'user.username') or simple fields
                val = obj
                for attr in field.split('.'):
                    val = getattr(val, attr, '')
                    if val is None: break
            
            # Formatting
            if isinstance(val, (int, float)):
                pass # Default is fine
            else:
                val = str(val) if val is not None else ''
                
            ws.cell(row=row_num, column=col_num).value = val
            
    # Auto-adjust column width
    for column_cells in ws.columns:
        # Use first cell of the column to find the column letter
        col_letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2
        
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

def export_list_to_xlsx(data_list, filename, headers, field_keys):
    """
    Generates an XLSX file from a list of dictionaries.
    
    :param data_list: List of dictionaries to export.
    :param filename: Name of the file.
    :param headers: List of column headers.
    :param field_keys: List of keys in the dictionaries corresponding to headers.
    """
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse("Library 'openpyxl' not installed.", status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"
    
    # Header
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = openpyxl.styles.Font(bold=True)
        
    # Data
    for row_num, item in enumerate(data_list, 2):
        for col_num, key in enumerate(field_keys, 1):
            val = item.get(key, '')
            ws.cell(row=row_num, column=col_num).value = str(val) if val is not None else ''
            
    # Auto-adjust
    for column_cells in ws.columns:
        col_letter = get_column_letter(column_cells[0].column)
        max_len = 0
        for cell in column_cells:
            try:
                if cell.value: max_len = max(max_len, len(str(cell.value)))
            except: pass
        ws.column_dimensions[col_letter].width = max_len + 2
        
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
